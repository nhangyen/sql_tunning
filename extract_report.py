import openpyxl
from datetime import datetime
import os
import re

wb = openpyxl.load_workbook('baocao.xlsx')
landing = wb['100 BAO CAO']
checked = {}

# Today's date
today = datetime.today().strftime("%Y%m%d")
folder = "bk"
os.makedirs(folder, exist_ok=True)

for row in landing.iter_rows(min_row=1, max_row=2):
    cur_cell = row[6]  # Column G
    if cur_cell.value is None:
        continue
    # Extract sheet name from reference like 'sheet name'!A1
    match = re.match(r"'?(.*?)'?!", cur_cell.value)
    if not match:
        print(f"Skipping invalid reference: {cur_cell.value}")
        continue
    sheet_name = match.group(1)

    if checked.get(sheet_name) != 1:
        if sheet_name not in wb.sheetnames:
            print(f"Sheet {sheet_name} does not exist, skipping.")
            continue
        
        cur_sheet = wb[sheet_name]

        txt = "CREATE OR REPLACE "
        for sub_row in cur_sheet.iter_rows(min_row=1, max_row=cur_sheet.max_row):
            if sub_row[0].value is None:
                continue
            cell = sub_row[0]  # First column
            txt += str(cell.value) + "\n"

        # Safe filename with folder and date
        safe_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '_', '-'))
        filename = os.path.join(folder, f"{safe_name}_BK_{today}.sql")

        with open(filename, "w", encoding="utf-8") as f:
            f.write(txt)

        checked[sheet_name] = 1
        print(f"Written {filename}")

wb.close()
